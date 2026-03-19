"""Tests for RT reserves / ORDC data downloader (NP6-792-ER)."""

import io
import json
import sqlite3
import tempfile
import zipfile
from pathlib import Path
from unittest.mock import MagicMock, patch

import pandas as pd
import pytest
from openpyxl import Workbook

from prediction.src.data.ercot.reserves import (
    list_archive_files,
    download_archive,
    parse_xlsx,
    save_to_sqlite,
)


def _create_test_xlsx(path: Path, sheets: dict[str, list[dict]] | None = None):
    """Create a synthetic XLSX mimicking the NP6-792-ER format.

    Header is at row 8 (rows 0-7 are metadata/blank).
    """
    if sheets is None:
        sheets = {
            "Jan": [
                {
                    "SCED Timestamp": "2024-01-01 00:05:27",
                    "Repeated Hour Flag": "N",
                    "Batch ID": 1,
                    "System Lamda": 25.5,
                    "PRC": 5000.0,
                    "RTOLCAP": 4000.0,
                    "RTOFFCAP": 1000.0,
                    "RTORPA": 0.0,
                    "RTOFFPA": 0.0,
                    "RTOLHSL": 3500.0,
                    "RTBP": 30.0,
                    "RTORDPA": 0.5,
                },
                {
                    "SCED Timestamp": "2024-01-01 00:10:27",
                    "Repeated Hour Flag": "N",
                    "Batch ID": 2,
                    "System Lamda": 26.0,
                    "PRC": 4800.0,
                    "RTOLCAP": 3800.0,
                    "RTOFFCAP": 1000.0,
                    "RTORPA": 0.1,
                    "RTOFFPA": 0.0,
                    "RTOLHSL": 3400.0,
                    "RTBP": 31.0,
                    "RTORDPA": 0.6,
                },
            ],
            "Feb": [
                {
                    "SCED Timestamp": "2024-02-01 12:00:00",
                    "Repeated Hour Flag": "N",
                    "Batch ID": 1,
                    "System Lamda": 28.0,
                    "PRC": 6000.0,
                    "RTOLCAP": 5000.0,
                    "RTOFFCAP": 1000.0,
                    "RTORPA": 0.0,
                    "RTOFFPA": 0.0,
                    "RTOLHSL": 4500.0,
                    "RTBP": 29.0,
                    "RTORDPA": 0.0,
                },
            ],
        }

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    for sheet_name, rows in sheets.items():
        ws = wb.create_sheet(title=sheet_name)
        if not rows:
            continue

        # Write metadata in rows 1-8 (rows 0-7 in 0-indexed)
        ws.cell(row=1, column=1, value="Report: NP6-792-ER")
        ws.cell(row=2, column=1, value="Year: 2024")

        # Header at row 9 (row index 8 in 0-indexed → header=8 in pandas)
        headers = list(rows[0].keys())
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=9, column=col_idx, value=header)

        # Data rows starting at row 10
        for row_idx, row_data in enumerate(rows, 10):
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=row_idx, column=col_idx, value=row_data[header])

    wb.save(path)


class TestParseXlsx:
    def test_basic_parse(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            xlsx_path = Path(tmpdir) / "test.xlsx"
            _create_test_xlsx(xlsx_path)

            df = parse_xlsx(xlsx_path)
            assert len(df) == 3  # 2 from Jan + 1 from Feb
            assert "sced_timestamp" in df.columns
            assert "prc" in df.columns
            assert "rtordpa" in df.columns

    def test_timestamps_are_utc(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            xlsx_path = Path(tmpdir) / "test.xlsx"
            _create_test_xlsx(xlsx_path)

            df = parse_xlsx(xlsx_path)
            # Chicago is UTC-6 in January, so 00:05 CST → 06:05 UTC
            first_ts = df.iloc[0]["sced_timestamp"]
            assert "T06:05:27Z" in first_ts

    def test_empty_sheets_skipped(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            xlsx_path = Path(tmpdir) / "test.xlsx"
            _create_test_xlsx(xlsx_path, sheets={"Jan": [], "Feb": []})

            df = parse_xlsx(xlsx_path)
            assert df.empty

    def test_report_info_sheet_skipped(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            xlsx_path = Path(tmpdir) / "test.xlsx"
            sheets = {
                "Report Info": [],
                "Jan": [
                    {
                        "SCED Timestamp": "2024-01-01 00:05:27",
                        "Repeated Hour Flag": "N",
                        "Batch ID": 1,
                        "System Lamda": 25.5,
                        "PRC": 5000.0,
                        "RTOLCAP": 4000.0,
                        "RTOFFCAP": 1000.0,
                        "RTORPA": 0.0,
                        "RTOFFPA": 0.0,
                        "RTOLHSL": 3500.0,
                        "RTBP": 30.0,
                        "RTORDPA": 0.5,
                    },
                ],
            }
            _create_test_xlsx(xlsx_path, sheets=sheets)

            df = parse_xlsx(xlsx_path)
            assert len(df) == 1

    def test_column_names_cleaned(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            xlsx_path = Path(tmpdir) / "test.xlsx"
            _create_test_xlsx(xlsx_path)

            df = parse_xlsx(xlsx_path)
            # All columns should be lowercase with underscores
            for col in df.columns:
                assert col == col.lower()
                assert " " not in col


class TestSaveToSqlite:
    def _make_df(self):
        return pd.DataFrame([
            {
                "sced_timestamp": "2024-01-01T06:05:27Z",
                "repeated_hour": "N",
                "batch_id": 1,
                "system_lambda": 25.5,
                "prc": 5000.0,
                "rtolcap": 4000.0,
                "rtoffcap": 1000.0,
                "rtorpa": 0.0,
                "rtoffpa": 0.0,
                "rtolhsl": 3500.0,
                "rtbp": 30.0,
                "rtordpa": 0.5,
            },
            {
                "sced_timestamp": "2024-01-01T06:10:27Z",
                "repeated_hour": "N",
                "batch_id": 2,
                "system_lambda": 26.0,
                "prc": 4800.0,
                "rtolcap": 3800.0,
                "rtoffcap": 1000.0,
                "rtorpa": 0.1,
                "rtoffpa": 0.0,
                "rtolhsl": 3400.0,
                "rtbp": 31.0,
                "rtordpa": 0.6,
            },
        ])

    def test_roundtrip(self):
        df = self._make_df()
        with tempfile.TemporaryDirectory() as tmpdir:
            db_path = Path(tmpdir) / "test.db"
            save_to_sqlite(df, db_path)

            conn = sqlite3.connect(str(db_path))
            result = pd.read_sql("SELECT * FROM rt_reserves", conn)
            conn.close()

            assert len(result) == 2
            assert result.iloc[0]["prc"] == 5000.0

    def test_upsert(self):
        df = self._make_df()
        with tempfile.TemporaryDirectory() as tmpdir:
            db_path = Path(tmpdir) / "test.db"
            save_to_sqlite(df, db_path)
            save_to_sqlite(df, db_path)

            conn = sqlite3.connect(str(db_path))
            result = pd.read_sql("SELECT * FROM rt_reserves", conn)
            conn.close()

            assert len(result) == 2

    def test_empty_df_no_error(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            db_path = Path(tmpdir) / "test.db"
            save_to_sqlite(pd.DataFrame(), db_path)
            assert not db_path.exists()


class TestListArchiveFiles:
    def test_parses_response(self):
        mock_client = MagicMock()
        mock_response = MagicMock()
        mock_response.json.return_value = {
            "archives": [
                {
                    "docId": 12345,
                    "friendlyName": "HIST_RT_SCED_PRC_2024.zip",
                    "postDatetime": "2025-01-15T08:00:00",
                },
                {
                    "docId": 12346,
                    "friendlyName": "HIST_RT_SCED_PRC_2023.zip",
                    "postDatetime": "2024-01-10T08:00:00",
                },
            ]
        }
        mock_response.raise_for_status = MagicMock()
        mock_client.session.get.return_value = mock_response
        mock_client.get_headers.return_value = {"Authorization": "Bearer test"}

        result = list_archive_files(mock_client)
        assert len(result) == 2
        assert result[0]["docId"] == 12345
        assert "2024" in result[0]["friendlyName"]


class TestDownloadArchive:
    def test_extracts_xlsx(self):
        # Create a fake ZIP containing an XLSX
        xlsx_content = b"fake xlsx content"
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, "w") as zf:
            zf.writestr("test_data.xlsx", xlsx_content)
        zip_bytes = zip_buffer.getvalue()

        mock_client = MagicMock()
        mock_response = MagicMock()
        mock_response.content = zip_bytes
        mock_response.raise_for_status = MagicMock()
        mock_client.session.get.return_value = mock_response
        mock_client.get_headers.return_value = {"Authorization": "Bearer test"}

        with tempfile.TemporaryDirectory() as tmpdir:
            output_dir = Path(tmpdir)
            result = download_archive(mock_client, 12345, output_dir)

            assert result.name == "test_data.xlsx"
            assert result.exists()
            assert result.read_bytes() == xlsx_content

#!/usr/bin/env python3
"""Import ERCOT Fuel Mix data into SQLite fuel_mix_hist table."""

import sqlite3
import sys
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("pip install openpyxl"); sys.exit(1)

# Also handle old .xls files
try:
    import xlrd
    HAS_XLRD = True
except ImportError:
    HAS_XLRD = False

DB_PATH = Path(__file__).parent.parent / "data" / "ercot_archive.db"
FUEL_DIR = Path("/Users/Shared/transfer/raw/fuel_mix/extracted")

BATCH_SIZE = 10000


def init_table(conn):
    c = conn.cursor()
    c.execute("""
    CREATE TABLE IF NOT EXISTS fuel_mix_hist (
        delivery_date TEXT NOT NULL,
        fuel TEXT NOT NULL,
        settlement_type TEXT,
        interval_15min INTEGER NOT NULL,
        generation_mw REAL NOT NULL,
        PRIMARY KEY (delivery_date, fuel, interval_15min)
    )
    """)
    c.execute("CREATE INDEX IF NOT EXISTS idx_fm_date ON fuel_mix_hist(delivery_date)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_fm_fuel ON fuel_mix_hist(fuel)")
    conn.commit()


def import_xlsx(conn, filepath):
    """Import a single xlsx fuel mix file."""
    c = conn.cursor()
    count = 0

    try:
        wb = openpyxl.load_workbook(str(filepath), data_only=True)
    except Exception as e:
        print(f"  ERROR opening {filepath.name}: {e}")
        return 0

    import datetime as dt_module
    
    # Find month sheets (Jan, Feb, ... or Jan16, Feb16, ...)
    month_abbrs = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
    month_sheets = [s for s in wb.sheetnames 
                    if any(s.startswith(m) for m in month_abbrs)]

    for sheet_name in month_sheets:
        ws = wb[sheet_name]
        batch = []
        header_found = False
        interval_columns = []
        is_old_format = False  # Date-Fuel combined column

        for row in ws.iter_rows(values_only=True):
            if not header_found:
                if row and row[0]:
                    cell0 = str(row[0]).lower().strip()
                    if 'date' in cell0:
                        header_found = True
                        if 'fuel' in cell0:
                            # Old format: 'Date-Fuel', 'Total', time columns...
                            is_old_format = True
                            for i in range(2, len(row)):
                                if row[i] is not None:
                                    interval_columns.append(i)
                        else:
                            # New format: 'Date', 'Fuel', 'Settlement Type', 'Total', time columns...
                            for i in range(4, len(row)):
                                if row[i] is not None and (
                                    ':' in str(row[i]) or 
                                    isinstance(row[i], dt_module.time)):
                                    interval_columns.append(i)
                continue

            if not row or not row[0]:
                continue

            try:
                if is_old_format:
                    # Parse 'MM/DD/YY_FuelType'
                    cell0 = str(row[0]).strip()
                    if '_' not in cell0:
                        continue
                    date_part, fuel = cell0.split('_', 1)
                    try:
                        d = datetime.strptime(date_part, "%m/%d/%y")
                    except ValueError:
                        d = datetime.strptime(date_part, "%m/%d/%Y")
                    date_str = d.strftime("%Y-%m-%d")
                    settle_type = None
                else:
                    date_val = row[0]
                    if isinstance(date_val, datetime):
                        date_str = date_val.strftime("%Y-%m-%d")
                    else:
                        date_str = str(date_val).strip()[:10]
                    fuel = str(row[1]).strip()
                    settle_type = str(row[2]).strip() if len(row) > 2 and row[2] else None

                for idx, col_i in enumerate(interval_columns):
                    if col_i < len(row) and row[col_i] is not None:
                        try:
                            gen = float(row[col_i])
                            batch.append((date_str, fuel, settle_type, idx + 1, gen))
                        except (ValueError, TypeError):
                            continue

                if len(batch) >= BATCH_SIZE:
                    c.executemany("""
                    INSERT OR REPLACE INTO fuel_mix_hist
                    (delivery_date, fuel, settlement_type, interval_15min, generation_mw)
                    VALUES (?, ?, ?, ?, ?)
                    """, batch)
                    count += len(batch)
                    batch = []

            except Exception:
                continue

        if batch:
            c.executemany("""
            INSERT OR REPLACE INTO fuel_mix_hist
            (delivery_date, fuel, settlement_type, interval_15min, generation_mw)
            VALUES (?, ?, ?, ?, ?)
            """, batch)
            count += len(batch)

    wb.close()
    conn.commit()
    return count


def import_xls(conn, filepath):
    """Import old .xls fuel mix file using xlrd.
    
    Old format (2007-2016): 'Date-Fuel' combined column like '12/01/15_Biomass',
    time columns as Excel time fractions (0.0104=0:15, 0.0208=0:30, etc.)
    """
    if not HAS_XLRD:
        print(f"  SKIP {filepath.name} (xlrd not installed)")
        return 0

    c = conn.cursor()
    count = 0

    try:
        wb = xlrd.open_workbook(str(filepath))
    except Exception as e:
        print(f"  ERROR opening {filepath.name}: {e}")
        return 0

    # Skip summary sheets, find month sheets
    skip = {'summary', 'disclaimer'}

    for sheet_name in wb.sheet_names():
        if any(s in sheet_name.lower() for s in skip):
            continue

        ws = wb.sheet_by_name(sheet_name)
        if ws.nrows < 2:
            continue

        batch = []

        # First row is header
        header = [ws.cell_value(0, ci) for ci in range(ws.ncols)]

        # Detect format from header
        header0 = str(ws.cell_value(0, 0)).strip().lower()
        is_combined = 'fuel' in header0  # 'Date-Fuel' combined
        # For separate format: Date, Fuel, Total, intervals...
        # For combined format: Date-Fuel, Total, intervals...
        data_start_col = 2 if is_combined else 3

        # Count interval columns
        interval_cols = []
        for i in range(data_start_col, len(header)):
            if isinstance(header[i], float) and header[i] > 0:
                interval_cols.append(i)

        if not interval_cols:
            continue

        for row_idx in range(1, ws.nrows):
            try:
                if is_combined:
                    cell0 = str(ws.cell_value(row_idx, 0)).strip()
                    # Split on '_' or '-' after date part (MM/DD/YY)
                    import re as _re
                    m = _re.match(r'(\d{2}/\d{2}/\d{2,4})\s*[_\-]\s*(.*)', cell0)
                    if not m:
                        continue
                    date_part, fuel = m.group(1), m.group(2)
                    try:
                        dt = datetime.strptime(date_part, "%m/%d/%y")
                    except ValueError:
                        try:
                            dt = datetime.strptime(date_part, "%m/%d/%Y")
                        except ValueError:
                            continue
                    date_str = dt.strftime("%Y-%m-%d")
                else:
                    # Separate Date and Fuel columns; date is Excel serial number
                    date_val = ws.cell_value(row_idx, 0)
                    fuel = str(ws.cell_value(row_idx, 1)).strip()
                    if not fuel:
                        continue
                    if isinstance(date_val, float) and date_val > 30000:
                        date_tuple = xlrd.xldate_as_tuple(date_val, wb.datemode)
                        date_str = f"{date_tuple[0]:04d}-{date_tuple[1]:02d}-{date_tuple[2]:02d}"
                    elif isinstance(date_val, str):
                        try:
                            dt = datetime.strptime(date_val.strip(), "%m/%d/%y")
                            date_str = dt.strftime("%Y-%m-%d")
                        except ValueError:
                            continue
                    else:
                        continue

                for idx, col_i in enumerate(interval_cols):
                    val = ws.cell_value(row_idx, col_i)
                    if val is not None and val != '':
                        try:
                            gen = float(val)
                            batch.append((date_str, fuel, None, idx + 1, gen))
                        except (ValueError, TypeError):
                            continue

                if len(batch) >= BATCH_SIZE:
                    c.executemany("""
                    INSERT OR REPLACE INTO fuel_mix_hist
                    (delivery_date, fuel, settlement_type, interval_15min, generation_mw)
                    VALUES (?, ?, ?, ?, ?)
                    """, batch)
                    count += len(batch)
                    batch = []
            except Exception:
                continue

        if batch:
            c.executemany("""
            INSERT OR REPLACE INTO fuel_mix_hist
            (delivery_date, fuel, settlement_type, interval_15min, generation_mw)
            VALUES (?, ?, ?, ?, ?)
            """, batch)
            count += len(batch)

    wb.release_resources()
    conn.commit()
    return count


def main():
    print(f"Database: {DB_PATH}")
    conn = sqlite3.connect(str(DB_PATH))
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA synchronous=NORMAL")
    init_table(conn)

    files = sorted(FUEL_DIR.glob("*"))
    print(f"Found {len(files)} fuel mix files\n")

    total = 0
    for f in files:
        print(f"Processing: {f.name}")
        if f.suffix.lower() == '.xlsx':
            n = import_xlsx(conn, f)
        elif f.suffix.lower() == '.xls':
            n = import_xls(conn, f)
        else:
            print(f"  SKIP (unknown format)")
            continue
        print(f"  -> {n:,} records")
        total += n

    print(f"\nTotal: {total:,} records")

    c = conn.cursor()
    c.execute("SELECT COUNT(*), MIN(delivery_date), MAX(delivery_date), COUNT(DISTINCT fuel) FROM fuel_mix_hist")
    cnt, min_d, max_d, fuels = c.fetchone()
    print(f"fuel_mix_hist: {cnt:,} records, {min_d} to {max_d}, {fuels} fuel types")

    conn.close()


if __name__ == "__main__":
    main()

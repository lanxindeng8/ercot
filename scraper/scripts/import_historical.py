#!/usr/bin/env python3
"""
Import ERCOT historical data from raw zip/xlsx/csv files into SQLite.

Creates NEW tables (dam_lmp_hist, rtm_lmp_hist, dam_asmcpc_hist) so
existing cron job tables are not affected.

Data sources:
  - NP4-180-ER (DAMLZHBSPP): DAM Hub/Zone prices, xlsx, 2015-2026
  - NP6-785-ER (RTMLZHBSPP): RTM Hub/Zone prices, xlsx (15-min), 2015-2026
  - NP4-181-ER (DAMASMCPC): DAM Ancillary Service prices, csv, 2015-2026

Usage:
    python import_historical.py                    # Import all
    python import_historical.py --type dam         # DAM LMP only
    python import_historical.py --type rtm         # RTM LMP only
    python import_historical.py --type asmcpc      # DAM ASMCPC only
"""

import argparse
import csv
import io
import os
import re
import sqlite3
import sys
import zipfile
from datetime import datetime, timedelta
from pathlib import Path
from typing import List, Optional

# Try openpyxl
try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl required. pip install openpyxl")
    sys.exit(1)

DB_PATH = Path(__file__).parent.parent / "data" / "ercot_archive.db"

# Raw data directories (check both locations)
RAW_DIRS = [
    Path("/Users/Shared/transfer/raw"),
]

BATCH_SIZE = 50000


def find_files(pattern_dirs: list, extensions: list) -> list:
    """Find all matching files across raw directories."""
    files = []
    for base_dir in RAW_DIRS:
        for sub in pattern_dirs:
            d = base_dir / sub
            if not d.exists():
                continue
            for f in sorted(d.iterdir()):
                if f.is_file() and f.suffix.lower() in extensions:
                    files.append(f)
    return files


def extract_xlsx_from_zip(zip_path: Path, tmp_dir: Path) -> Optional[Path]:
    """Extract xlsx from zip, return path."""
    try:
        with zipfile.ZipFile(zip_path) as zf:
            xlsx_names = [n for n in zf.namelist() if n.endswith('.xlsx')]
            csv_names = [n for n in zf.namelist() if n.endswith('.csv')]
            if xlsx_names:
                zf.extract(xlsx_names[0], tmp_dir)
                return tmp_dir / xlsx_names[0]
            elif csv_names:
                zf.extract(csv_names[0], tmp_dir)
                return tmp_dir / csv_names[0]
    except Exception as e:
        print(f"  ERROR extracting {zip_path.name}: {e}")
    return None


def init_tables(conn: sqlite3.Connection):
    """Create historical tables."""
    c = conn.cursor()

    # DAM LMP historical (Hub/Zone only, hourly)
    c.execute("""
    CREATE TABLE IF NOT EXISTS dam_lmp_hist (
        delivery_date TEXT NOT NULL,
        hour_ending INTEGER NOT NULL,
        repeated_hour INTEGER NOT NULL DEFAULT 0,
        settlement_point TEXT NOT NULL,
        lmp REAL NOT NULL,
        PRIMARY KEY (delivery_date, hour_ending, repeated_hour, settlement_point)
    )
    """)
    c.execute("CREATE INDEX IF NOT EXISTS idx_dam_hist_date ON dam_lmp_hist(delivery_date)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_dam_hist_sp ON dam_lmp_hist(settlement_point)")

    # RTM LMP historical (Hub/Zone only, 15-min intervals)
    c.execute("""
    CREATE TABLE IF NOT EXISTS rtm_lmp_hist (
        delivery_date TEXT NOT NULL,
        delivery_hour INTEGER NOT NULL,
        delivery_interval INTEGER NOT NULL,
        repeated_hour INTEGER NOT NULL DEFAULT 0,
        settlement_point TEXT NOT NULL,
        settlement_point_type TEXT,
        lmp REAL NOT NULL,
        PRIMARY KEY (delivery_date, delivery_hour, delivery_interval, repeated_hour, settlement_point)
    )
    """)
    c.execute("CREATE INDEX IF NOT EXISTS idx_rtm_hist_date ON rtm_lmp_hist(delivery_date)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_rtm_hist_sp ON rtm_lmp_hist(settlement_point)")

    # DAM ASMCPC historical (Ancillary service clearing prices, hourly)
    c.execute("""
    CREATE TABLE IF NOT EXISTS dam_asmcpc_hist (
        delivery_date TEXT NOT NULL,
        hour_ending INTEGER NOT NULL,
        repeated_hour INTEGER NOT NULL DEFAULT 0,
        regdn REAL,
        regup REAL,
        rrs REAL,
        nspin REAL,
        ecrs REAL,
        PRIMARY KEY (delivery_date, hour_ending, repeated_hour)
    )
    """)
    c.execute("CREATE INDEX IF NOT EXISTS idx_asmcpc_hist_date ON dam_asmcpc_hist(delivery_date)")

    conn.commit()


def parse_hour(hour_str) -> int:
    """Parse hour ending from various formats: '01:00', '1', 1, '24:00'."""
    if isinstance(hour_str, (int, float)):
        return int(hour_str)
    s = str(hour_str).strip()
    if ':' in s:
        return int(s.split(':')[0])
    return int(s)


def parse_repeated_hour(flag) -> int:
    """Convert repeated hour flag to int: 'Y'/True -> 1, else 0."""
    if flag is None:
        return 0
    if isinstance(flag, bool):
        return 1 if flag else 0
    return 1 if str(flag).strip().upper() == 'Y' else 0


def normalize_date(date_val) -> str:
    """Normalize date to YYYY-MM-DD format."""
    if isinstance(date_val, datetime):
        return date_val.strftime("%Y-%m-%d")
    s = str(date_val).strip()
    # MM/DD/YYYY
    if re.match(r'\d{2}/\d{2}/\d{4}', s):
        return datetime.strptime(s, "%m/%d/%Y").strftime("%Y-%m-%d")
    # YYYY-MM-DD already
    if re.match(r'\d{4}-\d{2}-\d{2}', s):
        return s
    return s


def import_dam_lmp(conn: sqlite3.Connection):
    """Import DAM LMP xlsx files into dam_lmp_hist."""
    print("\n" + "=" * 60)
    print("Importing DAM LMP (NP4-180-ER / DAMLZHBSPP)")
    print("=" * 60)

    files = find_files(["dam_lmp", "NP4-180-ER", "daily_dam/NP4-180-ER"], [".zip", ".xlsx"])
    print(f"Found {len(files)} files")

    tmp_dir = Path("/tmp/ercot_import_dam")
    tmp_dir.mkdir(exist_ok=True)

    total = 0
    c = conn.cursor()

    for f in files:
        print(f"\nProcessing: {f.name}")

        if f.suffix.lower() == '.zip':
            xlsx_path = extract_xlsx_from_zip(f, tmp_dir)
            if not xlsx_path:
                continue
        else:
            xlsx_path = f

        try:
            wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
        except Exception as e:
            print(f"  ERROR opening: {e}")
            continue

        file_count = 0
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            batch = []
            header_found = False

            for row in ws.iter_rows(max_col=6, values_only=True):
                if not header_found:
                    if row and row[0] and 'date' in str(row[0]).lower():
                        header_found = True
                    continue

                if not row or not row[0] or len(row) < 5:
                    continue

                try:
                    date = normalize_date(row[0])
                    hour = parse_hour(row[1])
                    repeated = parse_repeated_hour(row[2])
                    sp = str(row[3]).strip()
                    price = float(row[4])

                    batch.append((date, hour, repeated, sp, price))

                    if len(batch) >= BATCH_SIZE:
                        c.executemany("""
                        INSERT OR REPLACE INTO dam_lmp_hist
                        (delivery_date, hour_ending, repeated_hour, settlement_point, lmp)
                        VALUES (?, ?, ?, ?, ?)
                        """, batch)
                        file_count += len(batch)
                        batch = []
                except Exception:
                    continue

            if batch:
                c.executemany("""
                INSERT OR REPLACE INTO dam_lmp_hist
                (delivery_date, hour_ending, repeated_hour, settlement_point, lmp)
                VALUES (?, ?, ?, ?, ?)
                """, batch)
                file_count += len(batch)

        wb.close()
        conn.commit()
        total += file_count
        print(f"  -> {file_count:,} records")

    print(f"\nDAM LMP total: {total:,} records imported")
    return total


def import_rtm_lmp(conn: sqlite3.Connection):
    """Import RTM LMP xlsx files into rtm_lmp_hist."""
    print("\n" + "=" * 60)
    print("Importing RTM LMP (NP6-785-ER / RTMLZHBSPP)")
    print("=" * 60)

    files = find_files(["rtm_lmp", "NP6-785-ER", "daily_dam/NP6-785-ER"], [".zip", ".xlsx"])
    print(f"Found {len(files)} files")

    tmp_dir = Path("/tmp/ercot_import_rtm")
    tmp_dir.mkdir(exist_ok=True)

    total = 0
    c = conn.cursor()

    for f in files:
        print(f"\nProcessing: {f.name}")

        if f.suffix.lower() == '.zip':
            extracted = extract_xlsx_from_zip(f, tmp_dir)
            if not extracted:
                continue
            xlsx_path = extracted
        else:
            xlsx_path = f

        try:
            wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
        except Exception as e:
            print(f"  ERROR opening: {e}")
            continue

        file_count = 0
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            batch = []
            header_found = False

            for row in ws.iter_rows(max_col=8, values_only=True):
                if not header_found:
                    if row and row[0] and 'date' in str(row[0]).lower():
                        header_found = True
                    continue

                if not row or not row[0] or len(row) < 7:
                    continue

                try:
                    date = normalize_date(row[0])
                    hour = int(row[1])
                    interval = int(row[2])
                    repeated = parse_repeated_hour(row[3])
                    sp = str(row[4]).strip()
                    sp_type = str(row[5]).strip() if row[5] else None
                    price = float(row[6])

                    batch.append((date, hour, interval, repeated, sp, sp_type, price))

                    if len(batch) >= BATCH_SIZE:
                        c.executemany("""
                        INSERT OR REPLACE INTO rtm_lmp_hist
                        (delivery_date, delivery_hour, delivery_interval, repeated_hour,
                         settlement_point, settlement_point_type, lmp)
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                        """, batch)
                        file_count += len(batch)
                        batch = []
                except Exception:
                    continue

            if batch:
                c.executemany("""
                INSERT OR REPLACE INTO rtm_lmp_hist
                (delivery_date, delivery_hour, delivery_interval, repeated_hour,
                 settlement_point, settlement_point_type, lmp)
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """, batch)
                file_count += len(batch)

        wb.close()
        conn.commit()
        total += file_count
        print(f"  -> {file_count:,} records")

    print(f"\nRTM LMP total: {total:,} records imported")
    return total


def import_dam_asmcpc(conn: sqlite3.Connection):
    """Import DAM ASMCPC csv files into dam_asmcpc_hist."""
    print("\n" + "=" * 60)
    print("Importing DAM ASMCPC (NP4-181-ER / DAMASMCPC)")
    print("=" * 60)

    files = find_files(["dam_asmcpc", "NP4-181-ER", "daily_dam/NP4-181-ER"], [".zip", ".csv"])
    print(f"Found {len(files)} files")

    tmp_dir = Path("/tmp/ercot_import_asmcpc")
    tmp_dir.mkdir(exist_ok=True)

    total = 0
    c = conn.cursor()

    for f in files:
        print(f"\nProcessing: {f.name}")

        if f.suffix.lower() == '.zip':
            extracted = extract_xlsx_from_zip(f, tmp_dir)  # also handles csv in zip
            if not extracted:
                continue
            csv_path = extracted
        else:
            csv_path = f

        try:
            with open(csv_path, 'r') as fh:
                reader = csv.reader(fh)
                header = next(reader)

                # Find column indices
                # Expected: Delivery Date, Hour Ending, Repeated Hour Flag, REGDN, REGUP, RRS, NSPIN, ECRS
                # ECRS may not exist in older years
                col_map = {}
                for i, h in enumerate(header):
                    h_clean = h.strip().upper()
                    if 'DATE' in h_clean:
                        col_map['date'] = i
                    elif 'HOUR' in h_clean and 'REPEAT' not in h_clean:
                        col_map['hour'] = i
                    elif 'REPEAT' in h_clean:
                        col_map['repeated'] = i
                    elif h_clean == 'REGDN':
                        col_map['regdn'] = i
                    elif h_clean.startswith('REGUP'):
                        col_map['regup'] = i
                    elif h_clean == 'RRS':
                        col_map['rrs'] = i
                    elif h_clean == 'NSPIN':
                        col_map['nspin'] = i
                    elif h_clean == 'ECRS':
                        col_map['ecrs'] = i

                batch = []
                file_count = 0

                for row in reader:
                    if not row or not row[0].strip():
                        continue
                    try:
                        date = normalize_date(row[col_map['date']])
                        hour = parse_hour(row[col_map['hour']])
                        rep_idx = col_map.get('repeated')
                        repeated = parse_repeated_hour(row[rep_idx] if rep_idx is not None and rep_idx < len(row) else 'N')

                        def safe_float(col_name):
                            idx = col_map.get(col_name)
                            if idx is None or idx >= len(row):
                                return None
                            v = str(row[idx]).strip()
                            return float(v) if v else None

                        regdn = safe_float('regdn')
                        regup = safe_float('regup')
                        rrs = safe_float('rrs')
                        nspin = safe_float('nspin')
                        ecrs = safe_float('ecrs')

                        batch.append((date, hour, repeated, regdn, regup, rrs, nspin, ecrs))

                        if len(batch) >= BATCH_SIZE:
                            c.executemany("""
                            INSERT OR REPLACE INTO dam_asmcpc_hist
                            (delivery_date, hour_ending, repeated_hour, regdn, regup, rrs, nspin, ecrs)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                            """, batch)
                            file_count += len(batch)
                            batch = []
                    except Exception as e:
                        continue

                if batch:
                    c.executemany("""
                    INSERT OR REPLACE INTO dam_asmcpc_hist
                    (delivery_date, hour_ending, repeated_hour, regdn, regup, rrs, nspin, ecrs)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    """, batch)
                    file_count += len(batch)

        except Exception as e:
            print(f"  ERROR: {e}")
            continue

        conn.commit()
        total += file_count
        print(f"  -> {file_count:,} records")

    print(f"\nDAM ASMCPC total: {total:,} records imported")
    return total


def print_summary(conn: sqlite3.Connection):
    """Print summary of historical tables."""
    c = conn.cursor()
    print("\n" + "=" * 60)
    print("IMPORT SUMMARY")
    print("=" * 60)

    for table in ['dam_lmp_hist', 'rtm_lmp_hist', 'dam_asmcpc_hist']:
        try:
            c.execute(f"SELECT COUNT(*), MIN(delivery_date), MAX(delivery_date) FROM {table}")
            count, min_d, max_d = c.fetchone()
            c.execute(f"SELECT COUNT(DISTINCT settlement_point) FROM {table}" if 'asmcpc' not in table
                      else f"SELECT '-' FROM {table} LIMIT 1")
            sp_count = c.fetchone()[0]
            print(f"\n  {table}:")
            print(f"    Records: {count:,}")
            print(f"    Date range: {min_d} to {max_d}")
            if 'asmcpc' not in table:
                print(f"    Settlement points: {sp_count}")
        except Exception as e:
            print(f"  {table}: {e}")


def main():
    parser = argparse.ArgumentParser(description="Import ERCOT historical data")
    parser.add_argument("--type", choices=["dam", "rtm", "asmcpc", "all"], default="all")
    args = parser.parse_args()

    print(f"Database: {DB_PATH}")
    DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(str(DB_PATH))
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA synchronous=NORMAL")

    init_tables(conn)

    if args.type in ("dam", "all"):
        import_dam_lmp(conn)
    if args.type in ("rtm", "all"):
        import_rtm_lmp(conn)
    if args.type in ("asmcpc", "all"):
        import_dam_asmcpc(conn)

    print_summary(conn)
    conn.close()
    print("\nDone!")


if __name__ == "__main__":
    main()
